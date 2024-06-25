import pandas as pd
import os
from itertools import combinations
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
import data_mapping


class trader_combined:
    def __init__(self):
        self.file_directory = r"C:\GNA\Data\Trader Data"
        pass

    def unmerge_columns(self):
        input_file = os.path.join(self.file_directory, 'Trader.xlsx')

        # Load the workbook and the first sheet
        wb = load_workbook(input_file)
        ws = wb.active

        # Create a DataFrame from the sheet
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0]  # Set the first row as column headers
        df = df[1:]  # Skip the header row

        # Ensure 'Start Time (HH:MM)' column is in string format
        if 'Start Time (HH:MM)' in df.columns:
            df['Start Time (HH:MM)'] = df['Start Time (HH:MM)'].astype(str)
            df['End Time (HH:MM)'] = df['End Time (HH:MM)'].astype(str)


        # Iterate over merged cells and forward fill only those cells
        for merge in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merge.bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell_value = ws.cell(row=min_row, column=min_col).value
                    df.iat[row - 2, col - 1] = cell_value  # row-2 because df is zero-indexed and has no header

        print(df.head())

        # Save the modified DataFrame to an Excel file
        output_file = os.path.join(self.file_directory, 'trader_final.xlsx')
        df.to_excel(output_file, index=False)
        print(f'File saved {output_file}')

    def use_ffill_method(self):
        input_file = os.path.join(self.file_directory, 'trader_final.xlsx')
        df = pd.read_excel(input_file)

        df.iloc[:, :6] = df.iloc[:, :6].ffill()
        print(df.head())

        # Save the modified DataFrame to an Excel file
        output_file = os.path.join(self.file_directory, 'trader_final1.xlsx')
        df.to_excel(output_file, index=False)
        print(f'File saved {output_file}')

    def remove_unwanted_characters(self):
        input_file = os.path.join(self.file_directory, 'trader_final1.xlsx')
        df = pd.read_excel(input_file)

        # Columns to be processed
        columns_to_strip = ['Start date (DD/MM/YYYY)', 'End date (DD/MM/YYYY)', 'Start Time (HH:MM)',
                            'End Time (HH:MM)', 'Scheduled Volume (Mus)', 'Purchase Price (Rs/Kwh)',
                            'Sale Price/Transaction Price (Rs/kwh)', 'Trading Margin (Rs/kwh)']

        # Remove unwanted characters
        unwanted_chars = [' ', '*', ',', '^', '#', 'NIL', '•',"'",';']
        for char in unwanted_chars:
            df[columns_to_strip] = df[columns_to_strip].astype(str).apply(lambda x: x.str.replace(char, ''))

        # Replace specific characters in date columns
        date_columns = ['Start date (DD/MM/YYYY)', 'End date (DD/MM/YYYY)']

        # Replace specific characters in date columns
        df[date_columns] = df[date_columns].astype(str).apply(lambda x: x.str.replace('.', '-'))
        df[date_columns] = df[date_columns].astype(str).apply(lambda x: x.str.replace('/', '-'))
        df[date_columns] = df[date_columns].astype(str).apply(lambda x: x.str.replace('l', '1'))
        df[date_columns] = df[date_columns].astype(str).apply(lambda x: x.str.replace('−', '-'))
        df[date_columns] = df[date_columns].astype(str).apply(lambda x: x.str.replace('O', '0'))
        df[date_columns] = df[date_columns].astype(str).apply(lambda x: x.str.replace('00:00:00', ''))

        for column in date_columns:
            # Function to convert date if in correct format
            def convert_date(date):
                try:
                    # Try to parse the date in the format '%Y-%m-%d'
                    parsed_date = pd.to_datetime(date, format='%Y-%m-%d', errors='raise')
                    # If successful, return the date in the format '%d-%m-%Y'
                    return parsed_date.strftime('%d-%m-%Y')
                except (ValueError, TypeError):
                    # If parsing fails, return the original value
                    return date

            # Apply the conversion function to each element in the column
            df[column] = df[column].apply(convert_date)

        month_mapping = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07',
                         'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12', 'Ju1': '07', 'Mär': '03',
                         '0ct': '10'}
        for abbr, num in month_mapping.items():
            df[date_columns] = df[date_columns].astype(str).apply(lambda x: x.str.replace(abbr, num))

        # Number mapping
        number_mapping = {'-1-': '-01-', '-2-': '-02-', '-3-': '-03-', '-4-': '-04-', '-5-': '-05-', '-6-': '-06-',
                          '-7-': '-07-', '-8-': '-08-', '-9-': '-09-'}
        for abbr, num in number_mapping.items():
            df[date_columns] = df[date_columns].astype(str).apply(lambda x: x.str.replace(abbr, num))

        float_columns = ['Scheduled Volume (Mus)', 'Purchase Price (Rs/Kwh)', 'Sale Price/Transaction Price (Rs/kwh)',
                         'Trading Margin (Rs/kwh)']

        # Remove unwanted characters
        unwanted_chars_2 = ['-', 'nan', 'DIV/0!', 'DependuponIEXrate', 'DepanduponIEX', 'IEXLinkedPrice', 'SolarPower',
                            '‐']
        for char in unwanted_chars_2:
            df[float_columns] = df[float_columns].astype(str).apply(lambda x: x.str.replace(char, ''))

        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('<.=', '<=')
        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('<.', '<=')
        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('>0.00', '')
        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('<=.', '<=')
        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('≤4Paise', '<=0.04')
        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('≤.', '<=')
        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('≤', '<=')
        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('TradingMarginofRs0.02/Kwh', '0.02')
        df['Trading Margin (Rs/kwh)'] = df['Trading Margin (Rs/kwh)'].str.replace('TradingMarginis<', '<=')

        print(df.head())

        # Save the modified DataFrame to an Excel file
        output_file = os.path.join(self.file_directory, 'trader_final2.xlsx')
        df.to_excel(output_file, index=False)
        print(f'File saved {output_file}')

    def correct_date_format(self):
        input_file = os.path.join(self.file_directory, 'trader_final2.xlsx')
        df = pd.read_excel(input_file)

        # Replace specific characters in date columns
        date_columns = ['Start date (DD/MM/YYYY)', 'End date (DD/MM/YYYY)']
        df[date_columns] = df[date_columns].astype(str)

        # Replace non-standard hyphens with standard hyphens
        df[date_columns] = df[date_columns].replace({'‐': '-'}, regex=True)

        df['start_date_1'] = df['Start date (DD/MM/YYYY)']
        df['end_date_1'] = df['End date (DD/MM/YYYY)']
        df['start_date_2'] = df['Start date (DD/MM/YYYY)']
        df['end_date_2'] = df['End date (DD/MM/YYYY)']

        date_column_1 = ['start_date_1', 'end_date_1']
        date_column_2 = ['start_date_2', 'end_date_2']

        # Convert date columns to datetime
        for column in date_column_1:
            df[column] = pd.to_datetime(df[column], format='%d-%m-%Y', errors='coerce').dt.date

        for column in date_column_2:
            df[column] = pd.to_datetime(df[column], format='%d-%m-%y', errors='coerce').dt.date

        for col_1, col_2 in zip(date_column_1, date_column_2):
            df[col_1].fillna(df[col_2], inplace=True)

        # Drop the temporary second set of columns
        df.drop(columns=date_column_2, inplace=True)

        # Rename the corrected columns to the original column names
        df.rename(columns={'start_date_1': 'Start date (DD/MM/YYYY)', 'end_date_1': 'End date (DD/MM/YYYY)'},
                  inplace=True)

        print(df.head())

        # Save the modified DataFrame to an Excel file
        output_file = os.path.join(self.file_directory, 'trader_final3.xlsx')
        df.to_excel(output_file, index=False)
        print(f'File saved {output_file}')

    def correcting_state_name(self):
        input_file = os.path.join(self.file_directory, 'trader_final3.xlsx')
        df = pd.read_excel(input_file)

        # List of columns containing state names
        state_columns = ['State of Seller', 'State of Buyer']

        # Extract unique state names from specified columns
        unique_states = df[state_columns].stack().unique()

        states_and_ut = ['Andaman and Nicobar Islands', 'Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar',
                         'Chandigarh', 'Chhattisgarh', 'Dadra and Nagar Haveli and Daman and Diu', 'Delhi', 'Goa',
                         'Gujarat', 'Haryana', 'Himachal Pradesh', 'Jammu and Kashmir', 'Jharkhand', 'Karnataka',
                         'Kerala', 'Ladakh', 'Lakshadweep', 'Madhya Pradesh', 'Maharashtra', 'Manipur', 'Meghalaya',
                         'Mizoram', 'Nagaland', 'Odisha', 'Puducherry', 'Punjab', 'Rajasthan', 'Sikkim', 'Tamil Nadu',
                         'Telangana', 'Tripura', 'Uttar Pradesh', 'Uttarakhand', 'West Bengal']

        # Initialize a list to store results
        results = []

        # Calculate fuzzy ratio for each unique state name
        for state in unique_states:
            best_ratio = 0
            best_match = ''
            for official_state in states_and_ut:
                ratio = fuzz.ratio(state, official_state)
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = official_state
            results.append({'Original Name': state, 'Matched Name': best_match, 'Fuzzy Ratio': best_ratio})

        # Create a DataFrame from results
        result_df = pd.DataFrame(results)

        # Save to Excel file
        output_file = os.path.join(self.file_directory, 'matched_states.xlsx')
        result_df.to_excel(output_file, index=False)
        return result_df

    def data_mapping_file(self):
        input_file = os.path.join(self.file_directory, 'trader_final3.xlsx')
        df = pd.read_excel(input_file)

        # List of columns containing state names
        state_columns = ['State of Seller', 'State of Buyer']

        for name, mane_to_change in data_mapping.trader_state_mapping.items():
            df[state_columns] = df[state_columns].replace(name, mane_to_change)

        # columns_to_remove_new_line = ['']

        df = df.apply(lambda x: str(x).replace('\n', ' ') if isinstance(x, str) else x)

        # Save the modified DataFrame to an Excel file
        output_file = os.path.join(self.file_directory, 'trader_final4.xlsx')
        df.to_excel(output_file, index=False)
        print(f'File saved {output_file}')

    def removing_duplicate_category_values(self):
        input_file = os.path.join(self.file_directory, 'trader_final4.xlsx')
        df = pd.read_excel(input_file)

        # List of columns containing category names
        category_columns = ['Category of Seller', 'Category of Buyer']

        for col in category_columns:
            df[col] = df[col].str.replace('\n', '')

        # Extract unique category names from specified columns
        unique_categories = df[category_columns].stack().unique()

        category_values = ['State Discom', 'Trader (Government)', 'Captive', 'Deemed Distribution Licensee', 'Consumer',
                           'Trader/OAC', 'DVVNL', 'Trust', 'Discom Consumer', 'Distribution Licensee', 'Bulk Consumer',
                           'Open Access Consumer', 'Generator', 'State Utility', 'Discom', 'Power Exchange', 'Utility',
                           'Industry', 'Industrial Consumer', 'Trader/Discom', 'District Licensee', 'Distribution',
                           'Industrial Buyer', 'Government', 'Hotel', 'Industry', 'Bulk Power', 'Commercial'
                                                                                                'ISGS',
                           'Government of Uttar Pradesh', 'Central Government Organisation', 'Thermal', 'RE Power',
                           'Hydro', 'IPP', 'Trader/OAC', 'Bagasse', 'Nodal Agency', 'OA Consumer', 'Biomass', 'IEX',
                           'Gas', 'Buyer', 'Non Solar', 'SPD', 'WPD', 'HPD', 'Cogen', 'Small Hydro', 'State Discom',
                           'ISGS', 'Trading Licensee', 'Solar', 'Discom Consumer', '']

        # Initialize a list to store results
        results = []

        # Calculate fuzzy ratio for each unique state name
        for state in unique_categories:
            best_ratio = 0
            best_match = ''
            for official_state in category_values:
                ratio = fuzz.ratio(state, official_state)
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = official_state
            results.append({'Original Name': state, 'Matched Name': best_match, 'Fuzzy Ratio': best_ratio})

        # Create a DataFrame from results
        result_df = pd.DataFrame(results)

        # Save to Excel file
        output_file = os.path.join(self.file_directory, 'matched_category.xlsx')
        result_df.to_excel(output_file, index=False)
        print(f'File saved {output_file}')

    def data_mapping_for_category(self):
        input_file = os.path.join(self.file_directory, 'trader_final4.xlsx')
        df = pd.read_excel(input_file)

        # List of columns containing state names
        category_columns = ['Category of Seller', 'Category of Buyer']

        for name, mane_to_change in data_mapping.trader_category_mapping.items():
            df[category_columns] = df[category_columns].replace(name, mane_to_change)

        # Save the modified DataFrame to an Excel file
        output_file = os.path.join(self.file_directory, 'trader_final5.xlsx')
        df.to_excel(output_file, index=False)
        print(f'File saved {output_file}')


if __name__ == "__main__":
    trader_data = trader_combined()
    trader_data.unmerge_columns()
    trader_data.use_ffill_method()
    trader_data.remove_unwanted_characters()
    trader_data.correct_date_format()
    trader_data.data_mapping_file()
    trader_data.data_mapping_for_category()
    pass
